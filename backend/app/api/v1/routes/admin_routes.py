"""管理员路由 - 教师账号管理"""
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import Response
from openpyxl import Workbook
from sqlalchemy.orm import Session
import io

from app.db.session import get_db
from app.core.security import require_role, get_password_hash
from app.core.response import success
from app.core.exceptions import BusinessException
from app.models.entities import (
    User, Course, Class, Material, Question, Announcement, AnnouncementClass,
    AnnouncementRead, TaskCompletion, PasswordResetRequest, Project, ProjectImage,
    ProjectLike, QuizAttempt, ShowcaseItem, ShowcaseItemImage, StudentClassEnrollment,
    StoredFile,
)
from app.schemas.common import AuthUser, CreateTeacherRequest, TeacherInfo, ResetRequestResolve
from app.services.auth_service import (
    get_reset_requests_for_admin,
    approve_reset_request,
    reject_reset_request,
    rotate_user_tokens,
)
from app.services.audit_service import create_audit_log, export_audit_logs, query_audit_logs
from app.services.soft_delete_service import list_deleted_resources, purge_resource, restore_resource, soft_delete

router = APIRouter()

DEFAULT_PASSWORD = "123456"


def _parse_audit_datetime(value: str | None):
    """解析审计日志日期参数，格式错误时返回业务错误。"""
    if not value:
        return None
    from datetime import datetime
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise BusinessException(400, "日期格式错误，请使用 ISO 日期时间格式") from exc


def _build_teacher_import_template() -> bytes:
    """生成教师批量导入模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "教师导入模板"
    ws.append(["姓名", "工号", "学院"])
    ws.append(["张老师", "T1001", "人工智能学院"])
    ws.append(["李老师", "T1002", "电子信息学院"])

    for column, width in {"A": 16, "B": 16, "C": 20}.items():
        ws.column_dimensions[column].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_teacher_info(user: User) -> dict:
    """构建教师信息字典"""
    return {
        "id": user.id,
        "name": user.name,
        "major": user.major or "",
        "created_at": user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
        "needs_password_change": user.needs_password_change,
    }


@router.get("/teachers", summary="获取教师列表", description="管理员：获取所有教师账号列表")
def list_teachers(
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    """获取所有教师列表"""
    teachers = db.query(User).filter(
        User.role == "teacher",
        User.deleted_at.is_(None),
    ).all()
    return success([_build_teacher_info(t) for t in teachers])


@router.post("/teachers", summary="创建教师账号", description="管理员：手动创建单个教师账号，默认密码 123456，首次登录需改密")
def create_teacher(
    req: CreateTeacherRequest,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    """手动创建单个教师账号，默认密码 123456，首次登录需改密"""
    existing = db.query(User).filter(User.id == req.id).first()
    if existing:
        raise BusinessException(400, f"工号 {req.id} 已存在")
    new_teacher = User(
        id=req.id,
        name=req.name,
        major=req.major or "",
        role="teacher",
        hashed_password=get_password_hash(DEFAULT_PASSWORD),
        needs_password_change=True,
    )
    db.add(new_teacher)
    create_audit_log(
        db,
        user=current_user,
        action="user.create",
        resource_type="users",
        resource_id=new_teacher.id,
        resource_name=new_teacher.name,
        details={"role": "teacher", "major": new_teacher.major},
    )
    db.commit()
    db.refresh(new_teacher)
    return success(_build_teacher_info(new_teacher))


@router.get("/teachers/import/template", summary="下载教师导入模板", description="管理员：下载教师 Excel 批量导入模板")
def download_teacher_import_template(
    _: AuthUser = Depends(require_role("admin")),
):
    content = _build_teacher_import_template()
    headers = {
        "Content-Disposition": 'attachment; filename="teacher-import-template.xlsx"',
    }
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/teachers/import", summary="批量导入教师", description="管理员：通过 Excel 批量导入教师账号（第一列=姓名，第二列=工号，第三列=学院可选，第一行为表头）")
async def import_teachers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    """通过 Excel 批量导入教师账号（第一列=姓名，第二列=工号，第三列=学院可选，第一行为表头）"""
    try:
        import openpyxl
    except ImportError:
        raise BusinessException(500, "服务器缺少 openpyxl 依赖，请联系管理员")

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise BusinessException(400, "请上传 .xlsx 格式的 Excel 文件")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise BusinessException(400, "Excel 文件格式错误，无法解析")

    created_list = []
    skipped_list = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    for row in rows:
        if not row or len(row) < 2:
            continue
        name = str(row[0]).strip() if row[0] is not None else ""
        teacher_id = str(row[1]).strip() if row[1] is not None else ""
        major = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not name or not teacher_id:
            continue
        existing = db.query(User).filter(User.id == teacher_id).first()
        if existing:
            skipped_list.append(teacher_id)
            continue
        new_teacher = User(
            id=teacher_id,
            name=name,
            major=major,
            role="teacher",
            hashed_password=get_password_hash(DEFAULT_PASSWORD),
            needs_password_change=True,
        )
        db.add(new_teacher)
        created_list.append(teacher_id)

    db.commit()

    return success({
        "created_count": len(created_list),
        "skipped_count": len(skipped_list),
        "created": created_list,
        "skipped": skipped_list,
        "message": f"成功导入 {len(created_list)} 个教师账号，跳过 {len(skipped_list)} 个（工号已存在）",
    })


@router.post("/teachers/{teacher_id}/reset-password", summary="重置教师密码", description="管理员：将教师密码重置为 123456 并标记需要重新修改密码")
def reset_teacher_password(
    teacher_id: str,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    """重置教师密码为 123456，并标记需要重新修改密码"""
    teacher = db.query(User).filter(
        User.id == teacher_id,
        User.role == "teacher",
        User.deleted_at.is_(None),
    ).first()
    if not teacher:
        raise BusinessException(404, "教师不存在")
    teacher.hashed_password = get_password_hash(DEFAULT_PASSWORD)
    teacher.needs_password_change = True
    rotate_user_tokens(teacher)
    create_audit_log(
        db,
        user=current_user,
        action="user.password_reset",
        resource_type="users",
        resource_id=teacher.id,
        resource_name=teacher.name,
        details={"target_role": "teacher"},
    )
    db.commit()
    return success({"message": "密码已重置为 123456，教师下次登录需修改密码"})


@router.get("/teachers/{teacher_id}/dependencies", summary="查询教师关联数据", description="管理员：查询教师名下的课程、班级、公告数量，用于删除前确认")
def get_teacher_dependencies(
    teacher_id: str,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    """查询教师名下的关联数据统计"""
    teacher = db.query(User).filter(
        User.id == teacher_id,
        User.role == "teacher",
        User.deleted_at.is_(None),
    ).first()
    if not teacher:
        raise BusinessException(404, "教师不存在")
    course_count = db.query(Course).filter(Course.created_by == teacher_id).count()
    class_count = db.query(Class).filter(Class.created_by == teacher_id).count()
    announcement_count = db.query(Announcement).filter(Announcement.teacher_id == teacher_id).count()
    project_count = db.query(Project).filter(Project.author_id == teacher_id).count()
    showcase_count = db.query(ShowcaseItem).filter(ShowcaseItem.created_by == teacher_id).count()
    return success({
        "course_count": course_count,
        "class_count": class_count,
        "announcement_count": announcement_count,
        "project_count": project_count,
        "showcase_count": showcase_count,
    })


@router.delete("/teachers/{teacher_id}", summary="删除教师账号", description="管理员：软删除指定教师账号")
def delete_teacher(
    teacher_id: str,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    """软删除活跃教师账号，保留全部关联事实。"""
    teacher = db.query(User).filter(
        User.id == teacher_id,
        User.role == "teacher",
        User.deleted_at.is_(None),
    ).first()
    if not teacher:
        raise BusinessException(404, "教师不存在")
    soft_delete(db, teacher, current_user, action="user.delete")
    db.commit()
    return success({"message": "删除成功"})


# ── 密码重置申请管理（管理员端）─────────────────────────────────────────

@router.get("/password-reset-requests", summary="获取密码重置申请", description="管理员：查看所有密码重置申请，可选 status 筛选")
def list_all_reset_requests(
    status: str | None = None,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success(get_reset_requests_for_admin(db, status))


@router.post("/password-reset-requests/{request_id}/approve", summary="审批密码重置", description="管理员：审批通过任意密码重置申请")
def admin_approve_request(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    req = db.query(PasswordResetRequest).filter(PasswordResetRequest.id == request_id).first()
    target_user_id = req.user_id if req else None
    try:
        result = approve_reset_request(db, request_id, current_user.id)
        create_audit_log(
            db,
            user=current_user,
            action="user.password_reset",
            resource_type="users",
            resource_id=target_user_id,
            resource_name=target_user_id,
            details={"request_id": request_id, "result": "approved"},
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return success(result)


@router.post("/password-reset-requests/{request_id}/reject", summary="驳回密码重置", description="管理员：驳回任意密码重置申请")
def admin_reject_request(
    request_id: int,
    data: ResetRequestResolve,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    req = db.query(PasswordResetRequest).filter(PasswordResetRequest.id == request_id).first()
    target_user_id = req.user_id if req else None
    try:
        result = reject_reset_request(db, request_id, current_user.id, data.reason)
        create_audit_log(
            db,
            user=current_user,
            action="user.password_reset",
            resource_type="users",
            resource_id=target_user_id,
            resource_name=target_user_id,
            details={"request_id": request_id, "result": "rejected", "reason": data.reason},
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return success(result)


# ── 回收站与审计日志 ─────────────────────────────────────────────────────

@router.get("/deleted/{resource_type}", summary="查看已删除数据", description="管理员：查看指定资源类型的回收站数据")
def get_deleted_resources(
    resource_type: str,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success(list_deleted_resources(db, resource_type, page, page_size))


@router.post("/restore/{resource_type}/{resource_id}", summary="恢复已删除数据", description="管理员：恢复回收站中的数据；脱钩资料须指定 target_course_id")
def post_restore_resource(
    resource_type: str,
    resource_id: str,
    body: dict | None = None,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    payload = body or {}
    target_course_id = payload.get("target_course_id")
    if target_course_id is not None:
        try:
            target_course_id = int(target_course_id)
        except (TypeError, ValueError):
            raise BusinessException(400, "目标课程 ID 无效")
    return success(
        restore_resource(
            db,
            resource_type,
            resource_id,
            current_user,
            target_course_id=target_course_id,
        )
    )


@router.delete("/purge/{resource_type}/{resource_id}", summary="彻底删除数据", description="管理员：彻底删除回收站中的数据")
def delete_purged_resource(
    resource_type: str,
    resource_id: str,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    return success(purge_resource(db, resource_type, resource_id, current_user))


@router.get("/audit-logs", summary="查询审计日志", description="管理员：按用户、动作、资源类型和时间筛选审计日志")
def get_audit_logs(
    user_id: str | None = None,
    action: str | None = None,
    resource_type: str | None = None,
    resource_id: str | None = None,
    status: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success(query_audit_logs(
        db,
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        status=status,
        start_date=_parse_audit_datetime(start_date),
        end_date=_parse_audit_datetime(end_date),
        page=page,
        page_size=page_size,
    ))


@router.get("/audit-logs/export", summary="导出审计日志", description="管理员：导出审计日志 Excel")
def export_audit_logs_excel(
    user_id: str | None = None,
    action: str | None = None,
    resource_type: str | None = None,
    resource_id: str | None = None,
    status: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    content, export_count = export_audit_logs(
        db,
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        status=status,
        start_date=_parse_audit_datetime(start_date),
        end_date=_parse_audit_datetime(end_date),
    )
    create_audit_log(
        db,
        user=current_user,
        action="audit_log.export",
        resource_type="audit_logs",
        resource_name="审计日志导出",
        details={
            "export_count": export_count,
            "导出条数": export_count,
            "筛选条件": {
                "用户ID": user_id,
                "动作": action,
                "资源类型": resource_type,
                "资源ID": resource_id,
                "状态": status,
                "开始时间": start_date,
                "结束时间": end_date,
            },
        },
    )
    db.commit()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="audit-logs.xlsx"'},
    )


@router.get("/users/{user_id}/audit-logs", summary="查询用户操作历史", description="管理员：查看指定用户的审计日志")
def get_user_audit_logs(
    user_id: str,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success(query_audit_logs(db, user_id=user_id, page=page, page_size=page_size))


@router.get("/resources/{resource_type}/{resource_id}/audit-logs", summary="查询资源操作历史", description="管理员：查看指定资源的审计日志")
def get_resource_audit_logs(
    resource_type: str,
    resource_id: str,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success(query_audit_logs(db, resource_type=resource_type, resource_id=resource_id, page=page, page_size=page_size))
