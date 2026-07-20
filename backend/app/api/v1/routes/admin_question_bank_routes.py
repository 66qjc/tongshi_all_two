"""管理员独立共享题库路由。"""
from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessException
from app.core.response import paginated_success, success
from app.core.security import require_role
from app.core.upload_validation import ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE, validate_upload
from app.db.session import get_db
from app.schemas.common import AdminQuestionBatchDelete, AdminQuestionCreate, AdminQuestionUpdate, AuthUser
from app.services import admin_question_bank_service as service

router = APIRouter(prefix="/question-bank", tags=["admin-question-bank"])


def _build_template(question_type: str = "all") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "共享题库导入模板"
    ws.append(["题型", "标签", "课程名称", "题干", "选项（选择题用 | 分隔）", "答案", "解析"])
    if question_type == "choice":
        ws.append(["choice", "人工智能,基础", "", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
    elif question_type == "fill":
        ws.append(["fill", "通识常识", "", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
    elif question_type == "multi_choice":
        ws.append(["multi_choice", "编程基础|多选", "", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    else:
        ws.append(["choice", "人工智能,基础", "", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
        ws.append(["fill", "通识常识", "", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
        ws.append(["multi_choice", "编程基础|多选", "", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.get("", summary="共享题库列表", description="管理员：全站活跃共享题列表，不依赖公共课程入口")
def list_question_bank(
    type: str | None = None,
    keyword: str | None = None,
    tag: str | None = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    items, total, safe_page, safe_page_size = service.list_questions(
        db,
        type_=type,
        keyword=keyword,
        tag=tag,
        page=page,
        page_size=page_size,
    )
    return paginated_success([service.format_question(q) for q in items], total, safe_page, safe_page_size)


@router.get("/tags", summary="共享题库标签列表", description="聚合活跃共享题标签，供管理员端新增/筛选下拉选用")
def list_question_bank_tags(
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success(service.list_question_tags(db))


@router.post("", summary="新增共享题", description="管理员：可选择活跃公共课作为挂载快照，也可创建独立题")
def create_question_bank_item(
    data: AdminQuestionCreate,
    mount_course_id: int | None = Query(default=None, description="可选：活跃公共课程挂载"),
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    question = service.create_question(
        db,
        data.model_dump(),
        current_user,
        mount_course_id=mount_course_id,
    )
    return success(service.format_question(question))


@router.put("/{question_id}", summary="编辑共享题", description="管理员：编辑全站共享题")
def update_question_bank_item(
    question_id: int,
    data: AdminQuestionUpdate,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    question = service.update_question(db, question_id, data.model_dump(exclude_unset=True), current_user)
    if not question:
        raise BusinessException(404, "题目不存在")
    return success(service.format_question(question))


@router.delete("/{question_id}", summary="删除共享题", description="管理员：软删除题目并移入回收站")
def delete_question_bank_item(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    if not service.delete_question(db, question_id, current_user):
        raise BusinessException(404, "题目不存在")
    return success()


@router.post("/batch-delete", summary="批量删除共享题", description="管理员：批量软删除题目")
def batch_delete_question_bank(
    data: AdminQuestionBatchDelete,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    return success(service.delete_questions(db, data.question_ids, current_user))


@router.get("/import/template", summary="下载共享题库导入模板")
def download_template(
    template_type: str = Query("all", pattern="^(all|choice|fill|multi_choice)$"),
    _: AuthUser = Depends(require_role("admin")),
):
    content = _build_template(template_type)
    filename_map = {
        "choice": "admin-question-bank-choice-template.xlsx",
        "fill": "admin-question-bank-fill-template.xlsx",
        "multi_choice": "admin-question-bank-multi-choice-template.xlsx",
        "all": "admin-question-bank-template.xlsx",
    }
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_map[template_type]}"'},
    )


@router.post("/import", summary="导入共享题库", description="管理员：Excel 批量导入；题型、标签、题干、答案必填，课程名称可选，行内课程名称优先")
def import_question_bank(
    file: UploadFile = File(...),
    mount_course_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    content = file.file.read()
    err = validate_upload(
        file.filename,
        len(content),
        allowed_extensions=ALLOWED_EXCEL_EXTENSIONS,
        max_size=MAX_EXCEL_SIZE,
    )
    if err:
        raise BusinessException(400, err)
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise BusinessException(400, "Excel 文件读取失败，请确认文件是 .xlsx/.xls 格式，且没有损坏或被加密")
    ws = wb.active
    if ws.max_row < 2:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请至少保留表头并填写一行题目")
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required_header_groups = [
        ("题型", ["题型", "type"]),
        ("标签", ["标签", "课程标签", "tags", "course_tags"]),
        ("题干", ["题干", "stem"]),
        ("答案", ["答案", "answer"]),
    ]
    missing_headers = [
        label for label, candidates in required_header_groups if not any(candidate in headers for candidate in candidates)
    ]
    if missing_headers:
        raise BusinessException(400, f"Excel 表头缺少：{', '.join(missing_headers)}。请下载题库导入模板后按模板填写")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        rows.append(item)
    if not rows:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请填写题目内容后再上传")
    return success(service.import_questions(db, rows, current_user, mount_course_id=mount_course_id))


@router.get("/contributions", summary="题库贡献记录", description="管理员：全站贡献聚合，无课程上下文显示独立题库")
def list_contributions(
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    items, total, safe_page, safe_page_size = service.list_contributions(db, page, page_size)
    return paginated_success(
        [service.format_contribution(item) for item in items],
        total,
        safe_page,
        safe_page_size,
    )
