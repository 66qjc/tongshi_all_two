"""Question routes"""
from __future__ import annotations

from io import BytesIO

from typing import Optional

from fastapi import APIRouter, Depends, UploadFile, File, Query
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.security import get_current_user, require_role, require_roles
from app.core.response import success, paginated_success
from app.core.exceptions import BusinessException
from app.core.upload_validation import validate_upload, ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE
from app.schemas.common import AuthUser, QuestionCreate, QuestionUpdate
from app.services.question_service import (
    list_questions, create_question, update_question,
    get_course_questions, import_questions_from_excel, can_view_course_questions,
)

router = APIRouter(prefix="/questions", tags=["questions"])


def _format_question(q, current_user_id: str | None = None, *, include_answer: bool = True):
    course = q.course
    creator = getattr(q, "creator", None)
    data = {
        "id": q.id,
        "type": q.type,
        "course_id": q.course_id,
        "course_name": course.name if course else "",
        "stem": q.stem,
        "options": q.options or [],
        "tags": q.tags or [],
        "source_question_id": q.source_question_id,
        # 历史同步痕迹字段；全站共享后不再表示“公共/私有题库”。
        "is_synced": bool(q.source_question_id),
        "created_by": q.created_by,
        "creator_name": creator.name if creator else None,
        "star_rating": q.star_rating if q.star_rating is not None else 3,
        # 全站共享题库：是否可编辑以题目创建人为准，不是课程创建人。
        "is_owner": bool(current_user_id and q.created_by == current_user_id),
    }
    # 学生拉题不得预下发标准答案；提交后由 /quiz/submit 返回对错与解析。
    if include_answer:
        data["answer"] = q.answer
        data["explanation"] = q.explanation or ""
    else:
        data["answer"] = ""
        data["explanation"] = ""
    return data


@router.get(
    "",
    summary="题目列表",
    description="全站共享题库列表：按题型和关键词筛选，支持分页。"
    " course_id 仅为兼容旧调用参数，不再按课程筛选。",
)
def get_questions(
    type: Optional[str] = None,
    course_id: Optional[int] = Query(
        default=None,
        description="兼容参数：历史客户端可能仍传课程 ID，服务端忽略，不按课程过滤",
    ),
    keyword: Optional[str] = None,
    tag: Optional[str] = Query(default=None, description="标签关键词，分页前过滤"),
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("teacher")),
):
    questions, total = list_questions(
        db,
        course_id,
        type,
        current_user.id,
        keyword,
        page,
        page_size,
        tag=tag,
    )
    return paginated_success([_format_question(q, current_user.id) for q in questions], total, page, page_size)


@router.get("/course/{course_id}", summary="课程题目", description="获取指定课程的题目（用于测验），可选 ?ids=1,2,3 过滤指定题目")
def get_course_questions_for_quiz(
    course_id: int,
    ids: str | None = None,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(get_current_user),
):
    if not can_view_course_questions(db, course_id, current_user.id, current_user.role):
        raise BusinessException(404, "课程不存在")
    # 学生自由练习仅返回其可见题（本课 + 公共共享题库）；教师/管理员仍看全站活跃题。
    student_id = current_user.id if current_user.role == "student" else None
    questions = get_course_questions(db, course_id, student_user_id=student_id)
    # 按 question_ids 过滤（作业入口传入）
    if ids:
        id_set = {int(i) for i in ids.split(",") if i.strip().isdigit()}
        questions = [q for q in questions if q.id in id_set]
    include_answer = current_user.role != "student"
    return success([
        _format_question(q, current_user.id, include_answer=include_answer)
        for q in questions
    ])


@router.post("", summary="新增题目", description="教师端：创建共享题；course_id 可选，不传则写入独立共享题（仅标签归属）")
def add_question(data: QuestionCreate, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    q = create_question(db, data.model_dump(), current_user.id)
    return success({"id": q.id})


@router.put("/{question_id}", summary="编辑题目", description="教师端：修改指定题目的内容")
def edit_question(question_id: int, data: QuestionUpdate, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    q = update_question(db, question_id, data.model_dump(exclude_unset=True), current_user.id)
    if not q:
        raise BusinessException(404, "题目不存在")
    return success()


@router.delete("/{question_id}", summary="删除题目", description="教师端：删除指定题目")
def remove_question(question_id: int, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    raise BusinessException(403, "教师不能删除题目，请联系管理员处理")


@router.post("/batch-delete", summary="批量删除题目", description="教师端：批量删除题目及其关联数据")


def batch_delete_questions(
    question_ids: list[int],
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("teacher")),
):
    raise BusinessException(403, "教师不能删除题目，请联系管理员处理")


def _build_question_template(question_type: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入模板"
    ws.append(["题型", "课程名称", "标签", "题干", "选项（选择题用 | 分隔）", "答案", "解析"])
    if question_type == "choice":
        ws.append(["choice", "", "人工智能基础", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
    elif question_type == "fill":
        ws.append(["fill", "", "通识常识", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
    elif question_type == "multi_choice":
        ws.append(["multi_choice", "", "编程基础", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    else:
        ws.append(["choice", "", "人工智能基础", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
        ws.append(["fill", "", "通识常识", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
        ws.append(["multi_choice", "", "编程基础", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _download_template_response(template_type: str):
    content = _build_question_template(template_type)
    filename_map = {
        "choice": "choice-question-template.xlsx",
        "fill": "fill-question-template.xlsx",
        "multi_choice": "multi-choice-question-template.xlsx",
        "all": "question-template.xlsx",
    }
    filename = filename_map[template_type]
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/import/template", summary="下载题目导入模板", description="下载 Excel 批量导入模板（中文表头，支持选择题和填空题）")
def download_question_template(template_type: str = Query("all", pattern="^(all|choice|fill|multi_choice)$"), current_user: AuthUser = Depends(require_roles("teacher", "admin"))):
    return _download_template_response(template_type)


@router.get("/import/template/choice", summary="下载选择题导入模板", description="下载选择题 Excel 模板")
def download_choice_question_template(current_user: AuthUser = Depends(require_roles("teacher", "admin"))):
    return _download_template_response("choice")


@router.get("/import/template/fill", summary="下载填空题导入模板", description="下载填空题 Excel 模板")
def download_fill_question_template(current_user: AuthUser = Depends(require_roles("teacher", "admin"))):
    return _download_template_response("fill")


@router.get("/import/template/multi_choice", summary="下载多选题导入模板", description="下载多选题 Excel 模板")
def download_multi_choice_question_template(current_user: AuthUser = Depends(require_roles("teacher", "admin"))):
    return _download_template_response("multi_choice")


@router.post("/import", summary="Excel 批量导入题目", description="教师端：上传 Excel 批量导入题目（.xlsx，题型、标签、题干、答案必填；课程名称可选）")
def import_questions(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: AuthUser = Depends(require_roles("teacher", "admin"))):
    content = file.file.read()
    err = validate_upload(file.filename, len(
        content), allowed_extensions=ALLOWED_EXCEL_EXTENSIONS, max_size=MAX_EXCEL_SIZE)
    if err:
        raise BusinessException(400, err)
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise BusinessException(400, "Excel 文件读取失败，请确认文件是 .xlsx/.xls 格式，且没有损坏或被加密")
    ws = wb.active
    if ws.max_row < 2:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请至少保留表头并填写一行题目")
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(
        ws.iter_rows(min_row=1, max_row=1))]
    required_header_groups = [
        ("题型", ["题型", "type"]),
        ("标签", ["标签", "课程标签", "tags", "course_tags"]),
        ("题干", ["题干", "stem"]),
        ("答案", ["答案", "answer"]),
    ]
    missing_headers = [
        label
        for label, candidates in required_header_groups
        if not any(candidate in headers for candidate in candidates)
    ]
    if missing_headers:
        raise BusinessException(400, f"Excel 表头缺少：{', '.join(missing_headers)}。请下载题库导入模板后按模板填写")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {headers[i]: row[i] if i < len(
            row) else None for i in range(len(headers))}
        rows.append(item)
    if not rows:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请填写题目内容后再上传")
    return success(import_questions_from_excel(db, rows, current_user.id, role=current_user.role))
