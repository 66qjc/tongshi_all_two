"""管理员公共课程路由。"""
from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessException
from app.core.response import paginated_success, success
from app.core.security import require_role
from app.core.upload_validation import ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE, validate_upload
from app.db.session import get_db
from app.models.entities import CourseStage, Material, Question
from app.schemas.common import (
    AdminMaterialUpdate,
    AdminPublicCourseCreate,
    AdminPublicCourseUpdate,
    AdminQuestionBatchDelete,
    AdminQuestionCreate,
    AdminQuestionUpdate,
    AuthUser,
    CourseStageCreate,
    CourseStageUpdate,
)
from app.services import admin_public_course_service as service
from app.services import admin_question_bank_service as question_bank_service
from app.services.course_stage_service import create_stage, delete_stage, format_stage_out, list_stages_for_course, update_stage
from app.services.public_course_sync_service import sync_stages_to_course_copies
from app.services.question_bank_service import count_all_questions
from app.services.soft_delete_service import soft_delete

router = APIRouter(prefix="/public-courses", tags=["admin-public-courses"])

# 旧公共课题库入口兼容：统一标记弃用并指向独立共享题库
_DEPRECATED_QUESTION_BANK_HEADERS = {
    "Deprecation": "true",
    "Link": '</api/admin/question-bank>; rel="successor-version"',
}


def _mark_question_bank_deprecated(response: Response) -> None:
    for key, value in _DEPRECATED_QUESTION_BANK_HEADERS.items():
        response.headers[key] = value


def _format_course(course, sync_info: dict | None = None, total_question_count: int = 0) -> dict:
    active_material_count = sum(material.deleted_at is None for material in course.materials)
    data = {
        "id": course.id,
        "name": course.name,
        "description": course.description or "",
        "created_at": course.created_at.isoformat() if course.created_at else "",
        "created_by": course.created_by,
        "is_public": bool(course.is_public),
        "material_count": active_material_count,
        # 全站共享题库：题目数必须由调用方传入全站活跃题数，禁止回落本课计数。
        "question_count": total_question_count,
    }
    if sync_info:
        data.update(sync_info)
    return data


def _format_material(material) -> dict:
    return {
        "id": material.id,
        "course_id": material.course_id,
        "course_name": material.course.name if material.course else "",
        "type": material.type,
        "title": material.title,
        "url": material.url,
        "duration": material.duration,
        "pages": material.pages,
        "size": material.size,
        "date": material.date,
        "file_id": material.file_id,
        "source_material_id": material.source_material_id,
        "is_synced": bool(material.source_material_id),
        "stage_id": material.stage_id,
    }


def _mount_course_state(question) -> str:
    """挂载课程状态：active / soft_deleted / purged / none。"""
    if question.course_id is None:
        snapshot = (getattr(question, "mount_course_name_snapshot", None) or "").strip()
        return "purged" if snapshot else "none"
    course = question.course
    if course is None:
        return "purged"
    if course.deleted_at is not None:
        return "soft_deleted"
    return "active"


def _format_question(question) -> dict:
    course = question.course
    creator = getattr(question, "creator", None)
    mount_state = _mount_course_state(question)
    snapshot = (getattr(question, "mount_course_name_snapshot", None) or "").strip()
    if mount_state == "soft_deleted" and course is not None:
        course_name = f"{course.name}（已删除）"
    elif mount_state == "purged":
        course_name = snapshot or "原挂载课程已清理"
    elif mount_state == "none":
        course_name = "独立题库"
    else:
        course_name = course.name if course else snapshot
    return {
        "id": question.id,
        "type": question.type,
        "course_id": question.course_id,
        "course_name": course_name,
        "mount_course_state": mount_state,
        "mount_course_name_snapshot": snapshot,
        "stem": question.stem,
        "options": question.options or [],
        "answer": question.answer,
        "explanation": question.explanation or "",
        "tags": question.tags or [],
        "source_question_id": question.source_question_id,
        "is_synced": bool(question.source_question_id),
        "created_by": question.created_by,
        "creator_name": creator.name if creator else None,
        "star_rating": question.star_rating if question.star_rating is not None else 3,
    }


def _format_contribution(log) -> dict:
    return {
        "id": log.id,
        "public_course_id": log.public_course_id,
        "public_course_name": log.public_course_name,
        "operator_id": log.operator_id,
        "operator_name": log.operator_name,
        "operator_role": log.operator_role,
        "action": log.action,
        "question_count": log.question_count,
        "created_at": log.created_at.isoformat() if log.created_at else "",
    }


@router.get("", summary="公共课程列表", description="管理员：获取所有公共课程，含同步状态摘要")
def get_public_courses(
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    courses = service.list_public_courses(db)
    result = []
    total_question_count = count_all_questions(db)
    for course in courses:
        sync_info = service.get_course_sync_status(db, course)
        result.append(_format_course(course, sync_info, total_question_count))
    return success(result)


@router.post("", summary="创建公共课程", description="管理员：创建公共课程")
def add_public_course(
    data: AdminPublicCourseCreate,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    course = service.create_public_course(db, data.name.strip(), current_user.id, data.description or "")
    return success(_format_course(course, total_question_count=count_all_questions(db)))


@router.put("/{course_id}", summary="编辑公共课程", description="管理员：修改公共课程名称并同步教师副本")
def edit_public_course(
    course_id: int,
    data: AdminPublicCourseUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    course = service.update_public_course(db, course_id, data.name.strip(), data.description)
    if not course:
        raise BusinessException(404, "公共课程不存在")
    return success(_format_course(course, total_question_count=count_all_questions(db)))


@router.delete("/{course_id}", summary="删除公共课程", description="管理员：软删除公共课程，不转挂题目")
def remove_public_course(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    if not service.delete_public_course(db, course_id, operator=current_user):
        raise BusinessException(404, "公共课程不存在")
    return success()


@router.get("/{course_id}/stages", summary="公共课程阶段列表", description="管理员：获取公共课程阶段/目录")
def get_public_stages(
    course_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    stages = list_stages_for_course(db, course_id)
    if stages is None:
        raise BusinessException(404, "公共课程不存在")
    return success([format_stage_out(stage) for stage in stages])


@router.post("/{course_id}/stages", summary="新增公共课程阶段", description="管理员：为公共课程创建阶段/目录并同步教师副本")
def add_public_stage(
    course_id: int,
    data: CourseStageCreate,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    stage = create_stage(db, course_id, data.name.strip(), data.sort_order)
    sync_stages_to_course_copies(db, course)
    db.commit()
    db.refresh(stage)
    return success(format_stage_out(stage))


@router.put("/{course_id}/stages/{stage_id}", summary="编辑公共课程阶段", description="管理员：修改阶段名称/排序并同步教师副本")
def edit_public_stage(
    course_id: int,
    stage_id: int,
    data: CourseStageUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    # 先验证阶段归属，避免 update_stage 内部 flush 后才发现 course_id 不匹配
    pre_check = db.query(CourseStage).filter(CourseStage.id == stage_id).first()
    if not pre_check or pre_check.course_id != course_id:
        raise BusinessException(404, "阶段不存在")
    name = data.name.strip() if data.name is not None else None
    stage = update_stage(db, stage_id, name, data.sort_order)
    if not stage:
        raise BusinessException(404, "阶段不存在")
    sync_stages_to_course_copies(db, course)
    db.commit()
    db.refresh(stage)
    return success(format_stage_out(stage))


@router.delete(
    "/{course_id}/stages/{stage_id}",
    summary="删除公共课程阶段",
    description="管理员：删除阶段并同步教师副本；cascade_materials=true 时级联软删阶段下活跃资料",
)
def remove_public_stage(
    course_id: int,
    stage_id: int,
    cascade_materials: bool = False,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    stage = db.query(CourseStage).filter(CourseStage.id == stage_id, CourseStage.course_id == course_id).first()
    if not stage:
        raise BusinessException(404, "阶段不存在")
    source_has_active_materials = db.query(Material).filter(
        Material.stage_id == stage_id,
        Material.deleted_at.is_(None),
    ).first() is not None
    if source_has_active_materials and not cascade_materials:
        raise BusinessException(400, "阶段下仍有资料，请先移出或删除资料，或确认级联删除")

    # 删除源阶段前，先清理教师副本中引用该阶段的副本阶段
    copy_stages = db.query(CourseStage).filter(CourseStage.source_stage_id == stage_id).all()
    for copy_stage in copy_stages:
        active_copy = [
            m for m in (copy_stage.materials or [])
            if getattr(m, "deleted_at", None) is None
        ]
        if active_copy:
            if cascade_materials:
                for material in list(active_copy):
                    if material.source_material_id is None:
                        # 教师在副本阶段自行新增的资料不随公共源删除，转为未分类。
                        material.stage_id = None
                    else:
                        soft_delete(db, material, current_user, action="material.delete")
                        material.stage_id = None
            else:
                # 将副本阶段下活跃资料设为未分类，避免级联阻塞
                db.query(Material).filter(
                    Material.stage_id == copy_stage.id,
                    Material.deleted_at.is_(None),
                ).update(
                    {Material.stage_id: None}, synchronize_session=False,
                )
        db.delete(copy_stage)
    if not delete_stage(
        db,
        stage_id,
        cascade_materials=cascade_materials,
        operator_id=current_user.id,
        operator_role="admin",
    ):
        raise BusinessException(400, "阶段删除失败")
    sync_stages_to_course_copies(db, course)
    db.commit()
    return success()


@router.get("/{course_id}/materials", summary="公共课程资料列表", description="管理员：获取公共课程资料")
def get_public_materials(
    course_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success([_format_material(material) for material in service.list_public_materials(db, course_id)])


@router.post("/{course_id}/materials", summary="新增公共课程资料", description="管理员：新增资料并同步教师副本")
def add_public_material(
    course_id: int,
    data: AdminMaterialUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    material = service.create_public_material(db, course_id, data.model_dump())
    return success(_format_material(material))


@router.put("/{course_id}/materials/{material_id}", summary="编辑公共课程资料", description="管理员：修改资料并同步教师副本")
def edit_public_material(
    course_id: int,
    material_id: int,
    data: AdminMaterialUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    material = service.update_public_material(material_id=material_id, db=db, data=data.model_dump(exclude_unset=True))
    if not material or material.course_id != course_id:
        raise BusinessException(404, "公共资料不存在")
    return success(_format_material(material))


@router.delete("/{course_id}/materials/{material_id}", summary="删除公共课程资料", description="管理员：软删除公共资料，保留文件与教师副本引用")
def remove_public_material(
    course_id: int,
    material_id: int,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    material = service.get_public_material(db, material_id)
    if not material or material.course_id != course_id:
        raise BusinessException(404, "公共资料不存在")
    if not service.delete_public_material(db, material_id, operator=current_user):
        raise BusinessException(404, "公共资料不存在")
    return success()


@router.get(
    "/{course_id}/questions",
    summary="公共课程题库列表（已弃用）",
    description="兼容旧客户端：委托共享题库，并返回弃用响应头；请改用 /api/admin/question-bank",
)
def get_public_questions(
    course_id: int,
    response: Response,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    # 保留旧响应形态（全量数组）；course_id 仅校验入口公共课存在
    return success([
        question_bank_service.format_question(question)
        for question in service.list_public_questions(db, course_id)
    ])


@router.get(
    "/{course_id}/question-contributions",
    summary="公共课程题库贡献记录（已弃用）",
    description="兼容旧客户端：返回该公共课贡献历史，并标记弃用；请改用 /api/admin/question-bank/contributions",
)
def get_public_question_contributions(
    course_id: int,
    response: Response,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    items, total = service.list_question_contributions(db, course_id, page, page_size)
    return paginated_success([_format_contribution(item) for item in items], total, page, page_size)


def _build_admin_question_template(question_type: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "公共题目导入模板"
    ws.append(["题型", "标签", "题干", "选项（选择题用 | 分隔）", "答案", "解析"])
    if question_type == "choice":
        ws.append(["choice", "人工智能,基础", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
    elif question_type == "fill":
        ws.append(["fill", "通识常识", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
    elif question_type == "multi_choice":
        ws.append(["multi_choice", "编程基础|多选", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    else:
        ws.append(["choice", "人工智能,基础", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
        ws.append(["fill", "通识常识", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
        ws.append(["multi_choice", "编程基础|多选", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.get(
    "/questions/import/template",
    summary="下载公共题目导入模板（已弃用）",
    description="兼容旧客户端：请改用 /api/admin/question-bank/import/template",
)
def download_public_question_template(
    response: Response,
    template_type: str = Query("all", pattern="^(all|choice|fill|multi_choice)$"),
    _: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    content = _build_admin_question_template(template_type)
    filename_map = {
        "choice": "admin-choice-question-template.xlsx",
        "fill": "admin-fill-question-template.xlsx",
        "multi_choice": "admin-multi-choice-question-template.xlsx",
        "all": "admin-question-template.xlsx",
    }
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_map[template_type]}"',
            **_DEPRECATED_QUESTION_BANK_HEADERS,
        },
    )


@router.post(
    "/{course_id}/questions",
    summary="新增共享题库题目（已弃用）",
    description="兼容旧客户端：委托 /api/admin/question-bank，并以当前公共课作为挂载",
)
def add_public_question(
    course_id: int,
    data: AdminQuestionCreate,
    response: Response,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    question = question_bank_service.create_question(
        db,
        data.model_dump(),
        current_user,
        mount_course_id=course_id,
    )
    return success(question_bank_service.format_question(question))


@router.post(
    "/{course_id}/questions/import",
    summary="Excel 批量导入共享题库题目（已弃用）",
    description="兼容旧客户端：委托独立共享题库导入，并以当前公共课作为挂载",
)
def import_public_questions(
    course_id: int,
    response: Response,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    content = file.file.read()
    err = validate_upload(file.filename, len(content), allowed_extensions=ALLOWED_EXCEL_EXTENSIONS, max_size=MAX_EXCEL_SIZE)
    if err:
        raise BusinessException(400, err)
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise BusinessException(400, "Excel 文件读取失败，请确认文件是 .xlsx/.xls 格式，且没有损坏或被加密")
    ws = wb.active
    if ws.max_row < 2:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请至少保留表头并填写一行题目")
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required_header_groups = [
        ("题型", ["题型", "type"]),
        ("标签", ["标签", "课程标签", "tags", "course_tags"]),
        ("题干", ["题干", "stem"]),
        ("答案", ["答案", "answer"]),
    ]
    missing_headers = [
        label for label, candidates in required_header_groups if not any(candidate in headers for candidate in candidates)
    ]
    if missing_headers:
        raise BusinessException(400, f"Excel 表头缺少：{', '.join(missing_headers)}。请下载题库导入模板后按模板填写")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        rows.append(item)
    if not rows:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请填写题目内容后再上传")
    return success(question_bank_service.import_questions(
        db,
        rows,
        current_user,
        mount_course_id=course_id,
    ))


@router.put(
    "/{course_id}/questions/{question_id}",
    summary="编辑共享题库题目（已弃用）",
    description="兼容旧客户端：委托独立共享题库更新接口",
)
def edit_public_question(
    course_id: int,
    question_id: int,
    data: AdminQuestionUpdate,
    response: Response,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    # course_id 仅用于确认当前管理入口是有效公共课程；共享题目可挂在任意课程（含已软删挂载）。
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    question = question_bank_service.update_question(
        db,
        question_id,
        data.model_dump(exclude_unset=True),
        current_user,
    )
    if not question:
        raise BusinessException(404, "公共题目不存在")
    return success(question_bank_service.format_question(question))


@router.delete(
    "/{course_id}/questions/{question_id}",
    summary="删除共享题库题目（已弃用）",
    description="兼容旧客户端：委托独立共享题库软删除",
)
def remove_public_question(
    course_id: int,
    question_id: int,
    response: Response,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    # course_id 仅用于确认当前管理入口存在；共享题库删除不再要求题目挂在该公共课。
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    if not question_bank_service.delete_question(db, question_id, current_user):
        raise BusinessException(404, "公共题目不存在")
    return success()


@router.post(
    "/{course_id}/questions/batch-delete",
    summary="批量删除共享题库题目（已弃用）",
    description="兼容旧客户端：委托独立共享题库批量软删除",
)
def batch_remove_public_questions(
    course_id: int,
    data: AdminQuestionBatchDelete,
    response: Response,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    _mark_question_bank_deprecated(response)
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    return success(question_bank_service.delete_questions(db, data.question_ids, current_user))
