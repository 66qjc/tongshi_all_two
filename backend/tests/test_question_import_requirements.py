"""共享题库导入与贡献记录验收回归。"""
from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
import pytest

from app.models.entities import Course, Question, QuestionContributionLog, User
from app.services.admin_public_course_service import import_questions_to_public_course
from app.services.question_service import import_questions_from_excel
from tests.conftest import auth_header


def _admin_token(client) -> str:
    response = client.post("/api/token", json={"id": "admin", "password": "Admin#2026"})
    assert response.status_code == 200
    return response.json()["data"]["access_token"]


def _build_excel(headers: list[str], rows: list[list[str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _upload_questions(client, token: str, content: bytes):
    return client.post(
        "/api/questions/import",
        headers=auth_header(token),
        files={
            "file": (
                "questions.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_teacher_creating_independent_question_writes_independent_contribution(
    client,
    db_session,
    teacher_token,
):
    """教师新增未挂课程的共享题时应记录独立题库贡献。"""
    response = client.post(
        "/api/questions",
        headers=auth_header(teacher_token),
        json={
            "type": "fill",
            "stem": "教师独立贡献题",
            "options": [],
            "answer": "答案",
            "explanation": "",
            "tags": ["导入验收"],
        },
    )

    assert response.json()["code"] == 0
    log = db_session.query(QuestionContributionLog).one()
    assert log.public_course_id is None
    assert log.public_course_name == "独立题库"
    assert log.operator_id == "T001"
    assert log.operator_role == "teacher"
    assert log.action == "create"
    assert log.question_count == 1


def test_import_without_course_column_uses_tags_and_writes_independent_batch_log(
    client,
    db_session,
    teacher_token,
):
    """课程列可缺省，填写标签的题目应导入独立共享题库。"""
    content = _build_excel(
        ["题型", "标签", "题干", "选项", "答案", "解析"],
        [["fill", "人工智能、基础|人工智能", "仅标签导入题", "", "答案", ""]],
    )

    response = _upload_questions(client, teacher_token, content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 1
    assert payload["data"]["fail_count"] == 0
    question = db_session.query(Question).filter(Question.stem == "仅标签导入题").one()
    assert question.course_id is None
    assert question.tags == ["人工智能", "基础"]
    log = db_session.query(QuestionContributionLog).one()
    assert log.public_course_id is None
    assert log.public_course_name == "独立题库"
    assert log.action == "import"
    assert log.question_count == 1


def test_import_rejects_request_without_tag_header(client, teacher_token):
    """标签列缺失必须在请求级拒绝，而不是把题目写入空标签。"""
    content = _build_excel(
        ["题型", "题干", "选项", "答案", "解析"],
        [["fill", "缺标签表头题", "", "答案", ""]],
    )

    response = _upload_questions(client, teacher_token, content)

    payload = response.json()
    assert payload["code"] == 400
    assert "标签" in payload["message"]


@pytest.mark.parametrize("missing_header", ["题型", "题干", "答案"])
def test_import_rejects_request_without_each_remaining_required_header(client, teacher_token, missing_header):
    """题型、题干、答案表头任一缺失时，整份导入都必须被拒绝。"""
    values = {
        "题型": "fill",
        "标签": "有效标签",
        "题干": "必填表头失败题",
        "答案": "答案",
    }
    headers = [header for header in values if header != missing_header]
    content = _build_excel(headers, [[values[header] for header in headers]])

    response = _upload_questions(client, teacher_token, content)

    payload = response.json()
    assert payload["code"] == 400
    assert missing_header in payload["message"]


@pytest.mark.parametrize("tag_header", ["标签", "课程标签", "tags", "course_tags"])
def test_teacher_import_accepts_supported_tag_headers(client, db_session, teacher_token, tag_header):
    """教师导入必须兼容所有约定标签表头。"""
    stem = f"教师标签表头别名题-{tag_header}"
    content = _build_excel(
        ["题型", tag_header, "题干", "答案"],
        [["fill", "人工智能、基础", stem, "答案"]],
    )

    response = _upload_questions(client, teacher_token, content)

    assert response.json()["data"]["success_count"] == 1
    question = db_session.query(Question).filter(Question.stem == stem).one()
    assert question.tags == ["人工智能", "基础"]


@pytest.mark.parametrize("course_header", ["课程", "course", "course_name", "课程名称（可选）"])
def test_teacher_import_accepts_supported_course_headers(client, db_session, teacher_token, course_header):
    """教师导入必须兼容所有约定课程表头。"""
    course = Course(
        name=f"教师课程表头别名-{course_header}",
        created_by="T001",
        is_public=False,
    )
    db_session.add(course)
    db_session.commit()
    db_session.refresh(course)
    stem = f"教师课程表头别名题-{course_header}"
    content = _build_excel(
        ["题型", course_header, "标签", "题干", "答案"],
        [["fill", course.name, "有效标签", stem, "答案"]],
    )

    response = _upload_questions(client, teacher_token, content)

    assert response.json()["data"]["success_count"] == 1
    question = db_session.query(Question).filter(Question.stem == stem).one()
    assert question.course_id == course.id


def test_import_reports_empty_tag_row_without_rolling_back_valid_rows(
    client,
    db_session,
    teacher_token,
):
    """单行标签为空应失败，同文件有标签的行仍可成功。"""
    content = _build_excel(
        ["题型", "标签", "题干", "选项", "答案", "解析"],
        [
            ["fill", "， | 、", "空标签失败题", "", "答案", ""],
            ["fill", "有效标签", "有效标签成功题", "", "答案", ""],
        ],
    )

    response = _upload_questions(client, teacher_token, content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 1
    assert payload["data"]["fail_count"] == 1
    assert payload["data"]["errors"][0]["row"] == 2
    assert "标签" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "空标签失败题").count() == 0
    assert db_session.query(Question).filter(Question.stem == "有效标签成功题").count() == 1


def test_import_reports_empty_answer_row_without_writing_question(
    client,
    db_session,
    teacher_token,
):
    """答案列存在但单元格为空时，该行不得写入题库。"""
    content = _build_excel(
        ["题型", "标签", "题干", "选项", "答案", "解析"],
        [["fill", "有效标签", "空答案失败题", "", "", ""]],
    )

    response = _upload_questions(client, teacher_token, content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "答案不能为空" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "空答案失败题").count() == 0


def test_import_rejects_unknown_course_row_without_downgrading_to_independent_question(
    client,
    db_session,
    teacher_token,
):
    """填写了不存在课程的行必须失败，不能按独立题静默导入。"""
    content = _build_excel(
        ["题型", "课程名称（可选）", "标签", "题干", "选项", "答案", "解析"],
        [["fill", "不存在的课程", "有效标签", "未知课程失败题", "", "答案", ""]],
    )

    response = _upload_questions(client, teacher_token, content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "未找到课程" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "未知课程失败题").count() == 0


def test_admin_import_rejects_ambiguous_course_name(client, db_session):
    """管理员按课程名称导入时不得任意选择不同教师的同名课程。"""
    duplicate_name = "管理员同名课程歧义"
    db_session.add_all([
        Course(name=duplicate_name, created_by="T001", is_public=False),
        Course(name=duplicate_name, created_by="T002", is_public=False),
    ])
    db_session.commit()

    content = _build_excel(
        ["题型", "课程名称", "标签", "题干", "选项", "答案", "解析"],
        [["fill", duplicate_name, "有效标签", "同名课程拒绝题", "", "答案", ""]],
    )

    response = _upload_questions(client, _admin_token(client), content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "不唯一" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "同名课程拒绝题").count() == 0


def test_legacy_public_course_import_rejects_empty_answer(db_session):
    """保留的公共课程导入服务也不得写入空答案题目。"""
    course = Course(name="兼容公共课程空答案", created_by="admin", is_public=True)
    db_session.add(course)
    db_session.commit()

    result = import_questions_to_public_course(
        db_session,
        course.id,
        [{"题型": "fill", "标签": "公共标签", "题干": "兼容空答案失败题", "答案": ""}],
        operator_id="admin",
    )

    assert result["success_count"] == 0
    assert result["fail_count"] == 1
    assert "答案不能为空" in result["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "兼容空答案失败题").count() == 0


def test_teacher_import_keeps_valid_rows_when_database_flush_fails(db_session):
    """单行数据库写入失败不能使同一文件中已成功的题目回滚。"""
    from sqlalchemy import event

    failed_stem = "教师导入数据库失败题"
    injected = False

    def inject_duplicate_user(session, _flush_context, _instances):
        nonlocal injected
        if injected or not any(
            isinstance(item, Question) and item.stem == failed_stem
            for item in session.new
        ):
            return
        injected = True
        session.add(User(
            id="T001",
            name="重复教师",
            hashed_password="hash",
            role="teacher",
        ))

    event.listen(db_session, "before_flush", inject_duplicate_user)
    try:
        result = import_questions_from_excel(
            db_session,
            [
                {"题型": "fill", "标签": "有效标签", "题干": "教师导入有效题", "答案": "有效答案"},
                {"题型": "fill", "标签": "有效标签", "题干": failed_stem, "答案": "失败答案"},
            ],
            "T001",
        )
    finally:
        event.remove(db_session, "before_flush", inject_duplicate_user)

    assert result["success_count"] == 1
    assert result["fail_count"] == 1
    assert db_session.query(Question).filter(Question.stem == "教师导入有效题").count() == 1
    assert db_session.query(Question).filter(Question.stem == failed_stem).count() == 0
    logs = db_session.query(QuestionContributionLog).all()
    assert len(logs) == 1
    assert logs[0].question_count == 1


def test_downloaded_teacher_template_round_trips_as_independent_questions(
    client,
    db_session,
    teacher_token,
):
    """教师模板的示例课程应为空，下载后无需修改即可回传导入。"""
    template_response = client.get(
        "/api/questions/import/template",
        headers=auth_header(teacher_token),
    )
    assert template_response.status_code == 200
    workbook = load_workbook(io.BytesIO(template_response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]][1] == "课程名称"
    assert all(row[1].value in {None, ""} for row in sheet.iter_rows(min_row=2))

    response = _upload_questions(client, teacher_token, template_response.content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 3
    assert payload["data"]["fail_count"] == 0
    log = db_session.query(QuestionContributionLog).one()
    assert log.public_course_id is None
    assert log.question_count == 3


def test_admin_independent_question_uses_shared_contribution_record(client, db_session):
    """管理员独立新增题也应留下独立题库贡献记录。"""
    admin_token = _admin_token(client)
    response = client.post(
        "/api/admin/question-bank",
        headers=auth_header(admin_token),
        json={
            "type": "fill",
            "stem": "管理员独立贡献题",
            "options": [],
            "answer": "答案",
            "explanation": "",
            "tags": ["管理"],
        },
    )

    assert response.json()["code"] == 0
    log = db_session.query(QuestionContributionLog).one()
    assert log.public_course_id is None
    assert log.public_course_name == "独立题库"
    assert log.operator_id == "admin"
    assert log.operator_role == "admin"
    assert log.action == "create"


def test_teacher_import_aggregates_public_and_independent_contributions_separately(
    client,
    db_session,
):
    """同一批导入需按公共课程和独立题库分别聚合贡献数量。"""
    admin_token = _admin_token(client)
    public_course = Course(name="导入聚合公共课", created_by="admin", is_public=True)
    db_session.add(public_course)
    db_session.commit()

    content = _build_excel(
        ["题型", "课程名称", "标签", "题干", "选项", "答案", "解析"],
        [
            ["fill", public_course.name, "公共", "公共聚合题一", "", "答案", ""],
            ["fill", public_course.name, "公共", "公共聚合题二", "", "答案", ""],
            ["fill", "", "独立", "独立聚合题", "", "答案", ""],
        ],
    )

    response = _upload_questions(client, admin_token, content)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 3
    logs = db_session.query(QuestionContributionLog).order_by(QuestionContributionLog.id).all()
    assert [(item.public_course_id, item.question_count) for item in logs] == [
        (public_course.id, 2),
        (None, 1),
    ]
