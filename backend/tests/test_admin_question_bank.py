"""管理员独立共享题库测试。"""
import io

from openpyxl import Workbook, load_workbook
import pytest

from app.models.entities import Course, Question, QuestionContributionLog, User
from tests.conftest import auth_header


def _admin_token(client) -> str:
    resp = client.post("/api/token", json={"id": "admin", "password": "Admin#2026"})
    return resp.json()["data"]["access_token"]


def _build_import_excel(headers: list[str], rows: list[list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_admin_question_bank_works_without_public_course(client, db_session):
    """没有任何活跃公共课程时，管理员仍可列表、新增空挂载题、编辑和软删除。"""
    admin_token = _admin_token(client)
    # 确保没有活跃公共课
    db_session.query(Course).filter(Course.is_public.is_(True)).update(
        {Course.deleted_at: Course.created_at},
        synchronize_session=False,
    )
    db_session.commit()

    listed = client.get("/api/admin/question-bank", headers=auth_header(admin_token)).json()
    assert listed["code"] == 0

    created = client.post(
        "/api/admin/question-bank",
        json={
            "type": "fill",
            "stem": "独立题库新增题",
            "options": [],
            "answer": "北京",
            "explanation": "",
            "tags": ["独立"],
            "star_rating": 4,
        },
        headers=auth_header(admin_token),
    ).json()
    assert created["code"] == 0
    question_id = created["data"]["id"]
    assert created["data"]["course_id"] is None
    assert created["data"]["mount_course_state"] in {"none", "purged"}

    updated = client.put(
        f"/api/admin/question-bank/{question_id}",
        json={
            "type": "fill",
            "stem": "独立题库编辑后",
            "options": [],
            "answer": "上海",
            "explanation": "更新",
            "tags": ["独立"],
            "star_rating": 5,
        },
        headers=auth_header(admin_token),
    ).json()
    assert updated["code"] == 0
    assert updated["data"]["stem"] == "独立题库编辑后"

    deleted = client.delete(
        f"/api/admin/question-bank/{question_id}",
        headers=auth_header(admin_token),
    ).json()
    assert deleted["code"] == 0
    row = db_session.get(Question, question_id)
    assert row is not None
    assert row.deleted_at is not None

    log = (
        db_session.query(QuestionContributionLog)
        .filter(QuestionContributionLog.public_course_name == "独立题库")
        .order_by(QuestionContributionLog.id.desc())
        .first()
    )
    assert log is not None
    assert log.public_course_id is None


def test_legacy_public_course_question_routes_have_deprecation_headers(client, db_session):
    """旧公共课题库路由仍可用，但必须返回弃用响应头。"""
    admin_token = _admin_token(client)
    public = Course(
        name="弃用头测试公共课",
        description="",
        created_by="admin",
        is_public=True,
    )
    db_session.add(public)
    db_session.commit()
    db_session.refresh(public)

    listed = client.get(
        f"/api/admin/public-courses/{public.id}/questions",
        headers=auth_header(admin_token),
    )
    assert listed.status_code == 200
    assert listed.headers.get("Deprecation") == "true"
    assert "/api/admin/question-bank" in (listed.headers.get("Link") or "")
    assert listed.json()["code"] == 0

    created = client.post(
        f"/api/admin/public-courses/{public.id}/questions",
        json={
            "type": "fill",
            "stem": "旧入口新增题",
            "options": [],
            "answer": "答案",
            "explanation": "",
            "tags": [],
            "star_rating": 3,
        },
        headers=auth_header(admin_token),
    )
    assert created.status_code == 200
    assert created.headers.get("Deprecation") == "true"
    body = created.json()
    assert body["code"] == 0
    assert body["data"]["course_id"] == public.id


def test_admin_independent_import_requires_tags_and_writes_independent_log(client, db_session):
    """管理员独立导入缺标签表头应拒绝，合法标签题应记录独立贡献。"""
    admin_token = _admin_token(client)
    missing_tag = _build_import_excel(
        ["题型", "题干", "答案"],
        [["fill", "管理员缺标签题", "答案"]],
    )
    rejected = client.post(
        "/api/admin/question-bank/import",
        files={"file": ("missing-tag.xlsx", missing_tag, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_header(admin_token),
    ).json()
    assert rejected["code"] == 400
    assert "标签" in rejected["message"]

    tagged = _build_import_excel(
        ["题型", "标签", "题干", "答案"],
        [["fill", "管理员导入标签", "管理员独立导入题", "答案"]],
    )
    accepted = client.post(
        "/api/admin/question-bank/import",
        files={"file": ("tagged.xlsx", tagged, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_header(admin_token),
    ).json()
    assert accepted["code"] == 0
    assert accepted["data"]["success_count"] == 1
    question = db_session.query(Question).filter(Question.stem == "管理员独立导入题").one()
    assert question.course_id is None
    log = db_session.query(QuestionContributionLog).one()
    assert log.public_course_id is None
    assert log.public_course_name == "独立题库"
    assert log.operator_id == "admin"
    assert log.action == "import"


def test_admin_independent_import_rejects_empty_answer_row(client, db_session):
    """管理员导入中答案列为空的行不得进入独立共享题库。"""
    admin_token = _admin_token(client)
    content = _build_import_excel(
        ["题型", "标签", "题干", "答案"],
        [["fill", "管理员标签", "管理员空答案失败题", ""]],
    )

    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "empty-answer.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "答案不能为空" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "管理员空答案失败题").count() == 0


def test_admin_import_keeps_valid_rows_when_database_flush_fails(client, db_session):
    """管理员导入的单行数据库错误不能回滚同批已成功题目。"""
    from sqlalchemy import event

    failed_stem = "管理员导入数据库失败题"
    injected = False

    def inject_duplicate_user(session, _flush_context, _instances):
        nonlocal injected
        if injected or not any(
            isinstance(item, Question) and item.stem == failed_stem
            for item in session.new
        ):
            return
        injected = True
        session.add(User(
            id="admin",
            name="重复管理员",
            hashed_password="hash",
            role="admin",
        ))

    event.listen(db_session, "before_flush", inject_duplicate_user)
    try:
        content = _build_import_excel(
            ["题型", "标签", "题干", "答案"],
            [
                ["fill", "管理员标签", "管理员导入有效题", "有效答案"],
                ["fill", "管理员标签", failed_stem, "失败答案"],
            ],
        )
        response = client.post(
            "/api/admin/question-bank/import",
            files={
                "file": (
                    "database-failure.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=auth_header(_admin_token(client)),
        )
    finally:
        event.remove(db_session, "before_flush", inject_duplicate_user)

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 1
    assert payload["data"]["fail_count"] == 1
    assert db_session.query(Question).filter(Question.stem == "管理员导入有效题").count() == 1
    assert db_session.query(Question).filter(Question.stem == failed_stem).count() == 0
    log = db_session.query(QuestionContributionLog).one()
    assert log.question_count == 1


def test_admin_import_skips_duplicate_without_marking_row_failed(client):
    """管理员导入遇到全站重复题时应跳过，而不是因事务残留变为失败。"""
    admin_token = _admin_token(client)
    stem = "管理员重复导入跳过题"
    created = client.post(
        "/api/admin/question-bank",
        json={
            "type": "fill",
            "stem": stem,
            "options": [],
            "answer": "答案",
            "explanation": "",
            "tags": ["管理员标签"],
        },
        headers=auth_header(admin_token),
    )
    assert created.json()["code"] == 0
    content = _build_import_excel(
        ["题型", "标签", "题干", "答案"],
        [["fill", "管理员标签", stem, "答案"]],
    )

    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "duplicate-question.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["skip_count"] == 1
    assert payload["data"]["fail_count"] == 0


def test_admin_import_empty_course_name_creates_independent_question(client, db_session):
    """课程名称列存在但单元格为空时，管理员导入应保留独立题语义。"""
    admin_token = _admin_token(client)
    content = _build_import_excel(
        ["题型", "标签", "课程名称", "题干", "答案"],
        [["fill", "管理员标签", "", "管理员空课程名独立题", "答案"]],
    )

    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "empty-course-name.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 1
    question = db_session.query(Question).filter(Question.stem == "管理员空课程名独立题").one()
    assert question.course_id is None


def test_admin_import_rejects_unknown_course_name(client, db_session):
    """Excel 中填写不存在的课程名时，不能静默降级为独立题。"""
    admin_token = _admin_token(client)
    content = _build_import_excel(
        ["题型", "标签", "课程名称", "题干", "答案"],
        [["fill", "管理员标签", "不存在的管理员公共课", "管理员未知课程失败题", "答案"]],
    )

    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "unknown-course.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "不存在的管理员公共课" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "管理员未知课程失败题").count() == 0


def test_admin_import_rejects_private_course_name(client, db_session):
    """Excel 中填写私有课程名时，不能把题目导入独立题库。"""
    admin_token = _admin_token(client)
    private_course = Course(
        name="管理员私有课程名导入测试",
        description="",
        created_by="admin",
        is_public=False,
    )
    db_session.add(private_course)
    db_session.commit()

    content = _build_import_excel(
        ["题型", "标签", "课程名称", "题干", "答案"],
        [["fill", "管理员标签", private_course.name, "管理员私有课程失败题", "答案"]],
    )
    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "private-course.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "活跃公共课程" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "管理员私有课程失败题").count() == 0


def test_admin_import_public_course_name_attaches_questions_and_aggregates_contribution(client, db_session):
    """同一公共课程的多行导入应挂载该课，并合并为一条贡献记录。"""
    admin_token = _admin_token(client)
    public_course = Course(
        name="管理员公共课程名导入测试",
        description="",
        created_by="admin",
        is_public=True,
    )
    db_session.add(public_course)
    db_session.commit()
    db_session.refresh(public_course)

    stems = ["管理员公共课程导入题一", "管理员公共课程导入题二"]
    content = _build_import_excel(
        ["题型", "标签", "课程名称", "题干", "答案"],
        [
            ["fill", "管理员标签", public_course.name, stems[0], "答案一"],
            ["fill", "管理员标签", public_course.name, stems[1], "答案二"],
        ],
    )
    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "public-course.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 2
    questions = db_session.query(Question).filter(Question.stem.in_(stems)).all()
    assert {question.course_id for question in questions} == {public_course.id}
    log = db_session.query(QuestionContributionLog).filter(
        QuestionContributionLog.public_course_id == public_course.id,
        QuestionContributionLog.action == "import",
    ).one()
    assert log.public_course_name == public_course.name
    assert log.question_count == 2


def test_admin_import_course_name_overrides_default_mount_and_blank_cell_uses_default(client, db_session):
    """行内课程名优先；空课程单元格才回退到 mount_course_id。"""
    admin_token = _admin_token(client)
    default_course = Course(
        name="管理员默认挂载课程测试",
        description="",
        created_by="admin",
        is_public=True,
    )
    named_course = Course(
        name="管理员行内挂载课程测试",
        description="",
        created_by="T001",
        is_public=True,
    )
    db_session.add_all([default_course, named_course])
    db_session.commit()
    db_session.refresh(default_course)
    db_session.refresh(named_course)

    named_stem = "管理员行内课程优先题"
    fallback_stem = "管理员默认课程回退题"
    content = _build_import_excel(
        ["题型", "标签", "课程名称", "题干", "答案"],
        [
            ["fill", "管理员标签", named_course.name, named_stem, "答案一"],
            ["fill", "管理员标签", "", fallback_stem, "答案二"],
        ],
    )
    response = client.post(
        "/api/admin/question-bank/import",
        params={"mount_course_id": default_course.id},
        files={
            "file": (
                "course-name-priority.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 2
    named_question = db_session.query(Question).filter(Question.stem == named_stem).one()
    fallback_question = db_session.query(Question).filter(Question.stem == fallback_stem).one()
    assert named_question.course_id == named_course.id
    assert fallback_question.course_id == default_course.id


def test_admin_import_rejects_ambiguous_public_course_name(client, db_session):
    """同名活跃公共课不能由 .first() 任意选中，必须作为行错误返回。"""
    admin_token = _admin_token(client)
    duplicate_name = "管理员同名公共课程导入测试"
    db_session.add_all([
        Course(name=duplicate_name, description="", created_by="T001", is_public=True),
        Course(name=duplicate_name, description="", created_by="T002", is_public=True),
    ])
    db_session.commit()

    content = _build_import_excel(
        ["题型", "标签", "课程名称", "题干", "答案"],
        [["fill", "管理员标签", duplicate_name, "管理员同名课程失败题", "答案"]],
    )
    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "ambiguous-course.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 0
    assert payload["data"]["fail_count"] == 1
    assert "多门活跃公共课程" in payload["data"]["errors"][0]["reason"]
    assert db_session.query(Question).filter(Question.stem == "管理员同名课程失败题").count() == 0


@pytest.mark.parametrize("course_header", ["课程名称", "课程", "course", "course_name", "课程名称（可选）"])
def test_admin_import_accepts_supported_course_name_headers(client, db_session, course_header):
    """管理员导入应兼容所有约定的课程名称列表头。"""
    admin_token = _admin_token(client)
    public_course = Course(
        name=f"管理员课程列别名测试-{course_header}",
        description="",
        created_by="admin",
        is_public=True,
    )
    db_session.add(public_course)
    db_session.commit()
    db_session.refresh(public_course)

    stem = f"管理员课程列别名题-{course_header}"
    content = _build_import_excel(
        ["题型", "标签", course_header, "题干", "答案"],
        [["fill", "管理员标签", public_course.name, stem, "答案"]],
    )
    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "course-header-alias.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 1
    question = db_session.query(Question).filter(Question.stem == stem).one()
    assert question.course_id == public_course.id


def test_admin_question_bank_template_includes_optional_empty_course_name(client):
    """管理员模板应提供可选课程名称列，示例行不能伪造课程上下文。"""
    admin_token = _admin_token(client)
    response = client.get(
        "/api/admin/question-bank/import/template",
        headers=auth_header(admin_token),
    )

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    assert "课程名称" in headers
    course_column = headers.index("课程名称") + 1
    assert all(
        worksheet.cell(row=row_index, column=course_column).value in (None, "")
        for row_index in range(2, worksheet.max_row + 1)
    )


def test_admin_question_bank_template_round_trips_as_independent_questions(client, db_session):
    """管理员模板课程列为空时应无需修改即可回传为独立题。"""
    admin_token = _admin_token(client)
    template = client.get(
        "/api/admin/question-bank/import/template",
        headers=auth_header(admin_token),
    )

    response = client.post(
        "/api/admin/question-bank/import",
        files={
            "file": (
                "admin-question-template.xlsx",
                template.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(admin_token),
    )

    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["success_count"] == 3
    questions = db_session.query(Question).filter(Question.stem.in_([
        "图灵测试由谁提出？",
        "中国的首都是哪里？",
        "以下哪些是编程语言？",
    ])).all()
    assert {question.course_id for question in questions} == {None}
    log = db_session.query(QuestionContributionLog).filter(
        QuestionContributionLog.operator_id == "admin",
        QuestionContributionLog.action == "import",
    ).one()
    assert log.public_course_id is None
    assert log.question_count == 3
