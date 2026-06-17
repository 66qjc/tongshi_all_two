import io

from openpyxl import Workbook, load_workbook

from app.models.entities import User
from tests.conftest import auth_header


def _admin_headers(client):
    resp = client.post("/api/token", json={"id": "admin", "password": "Admin#2026"})
    assert resp.status_code == 200
    token = resp.json()["data"]["access_token"]
    return auth_header(token)


def _xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_download_teacher_import_template(client):
    resp = client.get("/api/admin/teachers/import/template", headers=_admin_headers(client))

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "teacher-import-template.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "教师导入模板"
    assert [cell.value for cell in ws[1]] == ["姓名", "工号", "学院"]
    assert [cell.value for cell in ws[2]] == ["张老师", "T1001", "人工智能学院"]
    assert [cell.value for cell in ws[3]] == ["李老师", "T1002", "电子信息学院"]


def test_import_teachers_reads_optional_major_column(client, db_session):
    content = _xlsx_bytes([
        ["姓名", "工号", "学院"],
        ["王老师", "T9001", "智能科学学院"],
    ])

    resp = client.post(
        "/api/admin/teachers/import",
        headers=_admin_headers(client),
        files={
            "file": (
                "teachers.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert resp.status_code == 200
    assert resp.json()["data"]["created_count"] == 1
    teacher = db_session.query(User).filter(User.id == "T9001").one()
    assert teacher.name == "王老师"
    assert teacher.major == "智能科学学院"
    assert teacher.role == "teacher"
    assert teacher.needs_password_change is True


def test_import_teachers_keeps_two_column_format_compatible(client, db_session):
    content = _xlsx_bytes([
        ["姓名", "工号"],
        ["赵老师", "T9002"],
    ])

    resp = client.post(
        "/api/admin/teachers/import",
        headers=_admin_headers(client),
        files={
            "file": (
                "teachers.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert resp.status_code == 200
    assert resp.json()["data"]["created_count"] == 1
    teacher = db_session.query(User).filter(User.id == "T9002").one()
    assert teacher.name == "赵老师"
    assert teacher.major == ""
